"""Gestion du classeur Excel, en ajout pur.

Le classeur appartient a l'utilisateur. Le script n'y ajoute que des lignes,
a la suite de l'existant, et ne reecrit jamais ce qui s'y trouve deja :
les colonnes ajoutees a la main, les couleurs, les commentaires et les onglets
personnels sont preserves d'un run a l'autre.

Le tri est chronologique croissant : les sorties les plus recentes arrivent
en bas, comme un journal. C'est ce qui rend l'ajout en fin de feuille naturel
sans jamais avoir a deplacer une ligne existante.

Structure : un onglet "Toutes sorties" en tete, puis un onglet par provider.
Colonnes : Date de sortie | Provider | Slot
"""

from __future__ import annotations

import logging
import re
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from . import config

log = logging.getLogger(__name__)

HEADERS = [
    "Date de sortie",
    "Provider",
    "Slot",
    "RTP",
    "Volatilite",
    "Max Win",
    "Gain System",
    "Gain System (brut)",
    "Reels",
    "Rows",
    "Lignes/Ways",
]

# Champ du state associe a chaque colonne, a partir de la colonne D.
ATTRIBUTE_COLUMNS = [
    ("rtp", "0.00"),
    ("volatility", None),
    ("max_win", "#,##0"),
    ("mechanic", None),
    ("mechanic_raw", None),
    ("reels", "0"),
    ("rows", "0"),
    ("paylines", "#,##0"),
]
FIRST_ATTRIBUTE_COL = 4

# Marqueur des donnees absentes chez le fournisseur.
NA = "N/A"

GLOBAL_SHEET = "Toutes sorties"
DATE_FORMAT = "DD/MM/YYYY"

HEADER_FILL = PatternFill("solid", fgColor="1F2933")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
COLUMN_WIDTHS = [16, 26, 42, 9, 12, 11, 15, 24, 8, 8, 13]

# Caracteres interdits par Excel dans un nom d'onglet.
_ILLEGAL_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")


# --- Utilitaires -----------------------------------------------------------


def _sheet_name(provider: str, taken: set[str]) -> str:
    """Nom d'onglet valide : 31 caracteres max, sans caractere interdit, unique."""
    cleaned = _ILLEGAL_SHEET_CHARS.sub("-", provider).strip() or "Inconnu"
    candidate = cleaned[:31]

    if candidate.lower() not in taken:
        taken.add(candidate.lower())
        return candidate

    for index in range(2, 100):
        suffix = f" ({index})"
        candidate = f"{cleaned[: 31 - len(suffix)]}{suffix}"
        if candidate.lower() not in taken:
            taken.add(candidate.lower())
            return candidate

    raise ValueError(f"Impossible de nommer un onglet pour '{provider}'")


def _parse_date(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value
    try:
        return datetime.strptime(str(value)[:10], "%Y-%m-%d")
    except ValueError:
        return None


def _row_key(provider: str, name: str) -> str:
    """Identifiant d'une ligne, insensible a la casse et aux espaces."""
    return f"{str(provider).strip().lower()}||{str(name).strip().lower()}"


def _sort_key(entry: dict) -> tuple:
    """Tri chronologique croissant : le plus ancien en haut."""
    parsed = _parse_date(entry.get("release_date"))
    return (
        parsed.toordinal() if parsed else 0,
        entry.get("provider", ""),
        entry.get("name", ""),
    )


# --- Ecriture --------------------------------------------------------------


def _style_header(sheet: Worksheet) -> None:
    sheet.append(HEADERS)
    for column in range(1, len(HEADERS) + 1):
        cell = sheet.cell(row=1, column=column)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for index, width in enumerate(COLUMN_WIDTHS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    sheet.freeze_panes = "A2"
    sheet.row_dimensions[1].height = 20


def _append_rows(sheet: Worksheet, entries: list[dict]) -> int:
    """Ajoute les lignes a la suite, sans toucher a l'existant."""
    for entry in entries:
        parsed = _parse_date(entry.get("release_date"))
        row = [parsed, entry.get("provider", ""), entry.get("name", "")]
        row += [
            entry.get(field) if entry.get(field) is not None else NA
            for field, _ in ATTRIBUTE_COLUMNS
        ]
        sheet.append(row)

        index = sheet.max_row
        if parsed is not None:
            sheet.cell(row=index, column=1).number_format = DATE_FORMAT
        for offset, (field, fmt) in enumerate(ATTRIBUTE_COLUMNS):
            if fmt and entry.get(field) is not None:
                sheet.cell(row=index, column=FIRST_ATTRIBUTE_COL + offset).number_format = fmt
    return len(entries)


def _existing_keys(sheet: Worksheet) -> set[str]:
    """Lignes deja presentes dans une feuille, lues depuis les colonnes B et C."""
    keys: set[str] = set()
    for provider, name in sheet.iter_rows(
        min_row=2, min_col=2, max_col=3, values_only=True
    ):
        if provider or name:
            keys.add(_row_key(provider or "", name or ""))
    return keys


def _refresh_autofilter(sheet: Worksheet) -> None:
    if sheet.max_row > 1:
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{sheet.max_row}"


# --- API du module ---------------------------------------------------------


def rebuild(entries: list[dict]) -> None:
    """Reconstruit le classeur a zero. Reserve au bootstrap et a --rebuild.

    Ecrase tout, y compris les colonnes ajoutees a la main. N'est jamais
    appele automatiquement une fois le classeur cree.
    """
    workbook = Workbook()
    workbook.remove(workbook.active)

    ordered = sorted(entries, key=_sort_key)

    global_sheet = workbook.create_sheet(GLOBAL_SHEET)
    _style_header(global_sheet)
    _append_rows(global_sheet, ordered)
    _refresh_autofilter(global_sheet)

    by_provider: dict[str, list[dict]] = {}
    for entry in ordered:
        by_provider.setdefault(entry.get("provider", "Inconnu"), []).append(entry)

    taken = {GLOBAL_SHEET.lower()}
    for provider in sorted(by_provider, key=str.casefold):
        sheet = workbook.create_sheet(_sheet_name(provider, taken))
        _style_header(sheet)
        _append_rows(sheet, by_provider[provider])
        _refresh_autofilter(sheet)

    config.DATA_DIR.mkdir(parents=True, exist_ok=True)
    workbook.save(config.EXCEL_FILE)
    log.info(
        "Classeur reconstruit : %s lignes, %s onglets provider",
        len(ordered),
        len(by_provider),
    )


def _fill_missing_cells(sheet: Worksheet, values: dict[str, dict]) -> int:
    """Comble les caracteristiques restees vides ou marquees N/A.

    Une valeur saisie a la main n'est jamais ecrasee : seules les cellules
    vides et les N/A poses par le script sont mises a jour, lorsque
    slot.report finit par publier la donnee.
    """
    filled = 0
    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=len(HEADERS)):
        entry = values.get(_row_key(row[1].value or "", row[2].value or ""))
        if entry is None:
            continue
        for offset, (field, fmt) in enumerate(ATTRIBUTE_COLUMNS):
            index = FIRST_ATTRIBUTE_COL + offset - 1
            if index >= len(row):
                continue
            cell = row[index]
            if cell.value not in (None, "", NA):
                continue
            value = entry.get(field)
            if value is None:
                if cell.value in (None, ""):
                    cell.value = NA
                continue
            cell.value = value
            if fmt:
                cell.number_format = fmt
            filled += 1
    return filled


def _fill_missing_dates(sheet: Worksheet, dates: dict[str, object]) -> int:
    """Complete la colonne date des lignes ou elle est restee vide.

    Une cellule deja remplie n'est jamais touchee : si tu as saisi une date a
    la main, elle est prioritaire sur celle du provider.
    """
    filled = 0
    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=3):
        date_cell, provider_cell, name_cell = row[0], row[1], row[2]
        if date_cell.value not in (None, ""):
            continue
        parsed = dates.get(_row_key(provider_cell.value or "", name_cell.value or ""))
        if parsed is None:
            continue
        date_cell.value = parsed
        date_cell.number_format = DATE_FORMAT
        filled += 1
    return filled


def sync(entries: list[dict]) -> None:
    """Ajoute au classeur existant les lignes qui n'y sont pas encore.

    Rien n'est jamais reecrit ni deplace : les colonnes personnelles, la mise
    en forme et les onglets ajoutes a la main survivent a chaque run.
    """
    if not config.EXCEL_FILE.exists():
        log.info("Classeur absent : reconstruction complete")
        rebuild(entries)
        return

    workbook = load_workbook(config.EXCEL_FILE)
    ordered = sorted(entries, key=_sort_key)

    # Dates devenues disponibles, a reporter sur les lignes restees vides.
    known_dates: dict[str, object] = {}
    for entry in ordered:
        parsed = _parse_date(entry.get("release_date"))
        if parsed is not None:
            known_dates[_row_key(entry.get("provider", ""), entry.get("name", ""))] = parsed

    by_key = {
        _row_key(e.get("provider", ""), e.get("name", "")): e for e in ordered
    }
    dates_filled = 0
    attrs_filled = 0

    # --- Onglet global -----------------------------------------------------
    if GLOBAL_SHEET in workbook.sheetnames:
        global_sheet = workbook[GLOBAL_SHEET]
    else:
        global_sheet = workbook.create_sheet(GLOBAL_SHEET, 0)
        _style_header(global_sheet)

    dates_filled += _fill_missing_dates(global_sheet, known_dates)
    attrs_filled += _fill_missing_cells(global_sheet, by_key)

    known = _existing_keys(global_sheet)
    missing = [e for e in ordered if _row_key(e.get("provider", ""), e.get("name", "")) not in known]
    added_global = _append_rows(global_sheet, missing)
    _refresh_autofilter(global_sheet)

    # --- Onglets par provider ---------------------------------------------
    by_provider: dict[str, list[dict]] = {}
    for entry in ordered:
        by_provider.setdefault(entry.get("provider", "Inconnu"), []).append(entry)

    taken = {name.lower() for name in workbook.sheetnames}
    # Correspondance provider -> onglet existant, etablie sur le nom tronque.
    added_provider = 0
    new_sheets = 0

    for provider, rows in by_provider.items():
        sheet = _find_provider_sheet(workbook, provider)
        if sheet is None:
            sheet = workbook.create_sheet(_sheet_name(provider, taken))
            _style_header(sheet)
            new_sheets += 1

        dates_filled += _fill_missing_dates(sheet, known_dates)
        _fill_missing_cells(sheet, by_key)
        known = _existing_keys(sheet)
        missing = [
            r for r in rows if _row_key(r.get("provider", ""), r.get("name", "")) not in known
        ]
        added_provider += _append_rows(sheet, missing)
        _refresh_autofilter(sheet)

    workbook.save(config.EXCEL_FILE)
    log.info(
        "Classeur mis a jour : %s ligne(s) ajoutee(s), %s date(s) et "
        "%s caracteristique(s) completees, %s nouvel(s) onglet(s)",
        added_global,
        dates_filled,
        attrs_filled,
        new_sheets,
    )


def _find_provider_sheet(workbook, provider: str) -> Worksheet | None:
    """Retrouve l'onglet d'un provider, en tolerant la troncature a 31 caracteres."""
    cleaned = _ILLEGAL_SHEET_CHARS.sub("-", provider).strip() or "Inconnu"
    for candidate in (cleaned, cleaned[:31]):
        for name in workbook.sheetnames:
            if name.lower() == candidate.lower():
                return workbook[name]
    return None
